from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.config.database import get_db
from app.api.routes.dependencies import get_current_user
from app.models.user import User
import openpyxl
from datetime import datetime
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


def _bulk_insert(db, table_name: str, columns: list[str], records: list[dict]):
    """Insert records in multi-row batches using raw SQLite literals to avoid D1 parameter limits."""
    if not records:
        return
    col_str = ", ".join(columns)
    val_rows = []
    for r in records:
        row_vals = []
        for k in columns:
            val = r.get(k)
            if val is None:
                row_vals.append("NULL")
            elif isinstance(val, (int, float)):
                row_vals.append(str(val))
            elif isinstance(val, bool):
                row_vals.append("1" if val else "0")
            else:
                escaped = str(val).replace("'", "''")
                row_vals.append(f"'{escaped}'")
        val_rows.append("(" + ", ".join(row_vals) + ")")
        
    # Batch execute in chunks of 50 to keep query sizes under D1 SQLITE_TOOBIG limits
    chunk_size = 50
    for i in range(0, len(val_rows), chunk_size):
        chunk = val_rows[i : i + chunk_size]
        sql = f"INSERT OR REPLACE INTO {table_name} ({col_str}) VALUES " + ", ".join(chunk)
        db.execute(text(sql))
    db.commit()

@router.get("/monthly/{month}")
async def get_monthly_report(month: str, db: Session = Depends(get_db)):
    """Get monthly report"""
    return {"report": {}}

@router.get("/mis-dashboard")
async def get_mis_dashboard_data(
    zone: str = None,
    district: str = None,
    coordinator: str = None,
    month: str = None,
    equipment: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Returns advanced operational analytics and BI intelligence from rj_penalties table.
    Supports dynamic multi-dimensional filtering and row-level access control.
    """
    try:
        from app.utils import cache
        cache_key = f"mis_dashboard:{current_user.user_id}:{zone}:{district}:{coordinator}:{month}:{equipment}"
        cached_val = cache.get(cache_key)
        if cached_val is not None:
            return cached_val
            
        # Check if table exists
        table_exists = db.execute(text(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='rj_penalties'"
        )).fetchone()
        
        if not table_exists:
            return {
                "success": False,
                "message": "Penalty database not seeded yet.",
                "summary": {}
            }

        # Row-level access control restriction
        user_role = current_user.role.strip() if current_user.role else "Engineer"
        
        # Enforce Zonal Manager restriction
        if user_role == "Zonal Manager":
            zone = current_user.zone
        # Enforce Coordinator restriction
        elif user_role == "Coordinator":
            coordinator = current_user.name
        # Enforce Engineer restriction
        elif user_role == "Engineer":
            district = current_user.district

        # Sanitize zone input (Ajmer Zone -> Ajmer)
        if zone:
            zone = zone.replace(" Zone", "").strip()

        # 1. Build dynamic WHERE clause based on filters
        where_clauses = ["1=1"]
        params = {}
        
        if district:
            where_clauses.append("LOWER(district_name) = LOWER(:district)")
            params["district"] = district
        if coordinator:
            where_clauses.append("LOWER(coordinator_name) = LOWER(:coordinator)")
            params["coordinator"] = coordinator
        if month:
            where_clauses.append("month_text = :month")
            params["month"] = month
        if equipment:
            where_clauses.append("equipment_name = :equipment")
            params["equipment"] = equipment
        if zone:
            zone_sql = """
                CASE 
                    WHEN district_name IN ('Ajmer', 'Bhilwara', 'Nagaur', 'Tonk', 'Beawer', 'Kekri', 'Shahpura') THEN 'Ajmer'
                    WHEN district_name IN ('Jaipur', 'Alwar', 'Dausa', 'Jhunjhunu', 'Sikar', 'Dudu', 'Kotputli', 'Neem Ka Thana', 'Khairthal') THEN 'Jaipur'
                    WHEN district_name IN ('Jodhpur', 'Barmer', 'Jaisalmer', 'Jalore', 'Pali', 'Sirohi', 'Phalodi', 'Balotra', 'Sanchore') THEN 'Jodhpur'
                    WHEN district_name IN ('Bikaner', 'Churu', 'Hanumangarh', 'Sri Ganganagar', 'Ganganagar', 'Anupgarh') THEN 'Bikaner'
                    WHEN district_name IN ('Kota', 'Baran', 'Bundi', 'Jhalawar') THEN 'Kota'
                    WHEN district_name IN ('Udaipur', 'Banswara', 'Chittorgarh', 'Dungarpur', 'Rajsamand', 'Pratapgarh', 'Salumbar') THEN 'Udaipur'
                    ELSE 'Other'
                END
            """
            where_clauses.append(f"LOWER({zone_sql}) = LOWER(:zone)")
            params["zone"] = zone

        where_str = " AND ".join(where_clauses)

        # 2. Query dynamic dependent dropdown lists based on active sibling filters
        zone_sql_expr = """
            CASE 
                WHEN district_name IN ('Ajmer', 'Bhilwara', 'Nagaur', 'Tonk', 'Beawer', 'Kekri', 'Shahpura') THEN 'Ajmer'
                WHEN district_name IN ('Jaipur', 'Alwar', 'Dausa', 'Jhunjhunu', 'Sikar', 'Dudu', 'Kotputli', 'Neem Ka Thana', 'Khairthal') THEN 'Jaipur'
                WHEN district_name IN ('Jodhpur', 'Barmer', 'Jaisalmer', 'Jalore', 'Pali', 'Sirohi', 'Phalodi', 'Balotra', 'Sanchore') THEN 'Jodhpur'
                WHEN district_name IN ('Bikaner', 'Churu', 'Hanumangarh', 'Sri Ganganagar', 'Ganganagar', 'Anupgarh') THEN 'Bikaner'
                WHEN district_name IN ('Kota', 'Baran', 'Bundi', 'Jhalawar') THEN 'Kota'
                WHEN district_name IN ('Udaipur', 'Banswara', 'Chittorgarh', 'Dungarpur', 'Rajsamand', 'Pratapgarh', 'Salumbar') THEN 'Udaipur'
                ELSE 'Other'
            END
        """

        # Districts dependent query
        where_d = ["1=1"]
        params_d = {}
        if zone:
            where_d.append(f"LOWER({zone_sql_expr}) = LOWER(:zone)")
            params_d["zone"] = zone
        if coordinator:
            where_d.append("LOWER(coordinator_name) = LOWER(:coordinator)")
            params_d["coordinator"] = coordinator
        if month:
            where_d.append("month_text = :month")
            params_d["month"] = month
        if equipment:
            where_d.append("equipment_name = :equipment")
            params_d["equipment"] = equipment
            
        districts_list = [r[0] for r in db.execute(text(f"""
            SELECT DISTINCT district_name FROM rj_penalties 
            WHERE {" AND ".join(where_d)} AND district_name IS NOT NULL AND district_name != '' 
            ORDER BY district_name
        """), params_d).fetchall()]

        # Coordinators dependent query
        where_c = ["1=1"]
        params_c = {}
        if zone:
            where_c.append(f"LOWER({zone_sql_expr}) = LOWER(:zone)")
            params_c["zone"] = zone
        if district:
            where_c.append("LOWER(district_name) = LOWER(:district)")
            params_c["district"] = district
        if month:
            where_c.append("month_text = :month")
            params_c["month"] = month
        if equipment:
            where_c.append("equipment_name = :equipment")
            params_c["equipment"] = equipment
            
        coordinators_list = [r[0] for r in db.execute(text(f"""
            SELECT DISTINCT coordinator_name FROM rj_penalties 
            WHERE {" AND ".join(where_c)} AND coordinator_name IS NOT NULL AND coordinator_name != '' 
            ORDER BY coordinator_name
        """), params_c).fetchall()]

        # Months dependent query
        where_m = ["1=1"]
        params_m = {}
        if zone:
            where_m.append(f"LOWER({zone_sql_expr}) = LOWER(:zone)")
            params_m["zone"] = zone
        if district:
            where_m.append("LOWER(district_name) = LOWER(:district)")
            params_m["district"] = district
        if coordinator:
            where_m.append("LOWER(coordinator_name) = LOWER(:coordinator)")
            params_m["coordinator"] = coordinator
        if equipment:
            where_m.append("equipment_name = :equipment")
            params_m["equipment"] = equipment
            
        months_list = [r[0] for r in db.execute(text(f"""
            SELECT DISTINCT month_text FROM rj_penalties 
            WHERE {" AND ".join(where_m)} AND month_text IS NOT NULL AND month_text != '' 
            ORDER BY month_text
        """), params_m).fetchall()]

        # Equipments dependent query
        where_eq = ["1=1"]
        params_eq = {}
        if zone:
            where_eq.append(f"LOWER({zone_sql_expr}) = LOWER(:zone)")
            params_eq["zone"] = zone
        if district:
            where_eq.append("LOWER(district_name) = LOWER(:district)")
            params_eq["district"] = district
        if coordinator:
            where_eq.append("LOWER(coordinator_name) = LOWER(:coordinator)")
            params_eq["coordinator"] = coordinator
        if month:
            where_eq.append("month_text = :month")
            params_eq["month"] = month
            
        equipments_list = [r[0] for r in db.execute(text(f"""
            SELECT DISTINCT equipment_name FROM rj_penalties 
            WHERE {" AND ".join(where_eq)} AND equipment_name IS NOT NULL AND equipment_name != '' 
            ORDER BY equipment_name
        """), params_eq).fetchall()]

        # Zones dependent query
        where_z = ["1=1"]
        params_z = {}
        if district:
            where_z.append("LOWER(district_name) = LOWER(:district)")
            params_z["district"] = district
        if coordinator:
            where_z.append("LOWER(coordinator_name) = LOWER(:coordinator)")
            params_z["coordinator"] = coordinator
        if month:
            where_z.append("month_text = :month")
            params_z["month"] = month
        if equipment:
            where_z.append("equipment_name = :equipment")
            params_z["equipment"] = equipment
            
        zones_list = [r[0] for r in db.execute(text(f"""
            SELECT DISTINCT {zone_sql_expr} as zone_name FROM rj_penalties 
            WHERE {" AND ".join(where_z)} AND district_name IS NOT NULL AND district_name != '' 
            ORDER BY zone_name
        """), params_z).fetchall()]

        # 3. Aggregated Total Metrics
        totals = db.execute(text(f"""
            SELECT 
                COUNT(*) as total_calls,
                SUM(CASE WHEN complaint_status = 'Final Closed' OR status = 'Closed' THEN 1 ELSE 0 END) as closed_calls,
                SUM(CASE WHEN is_ftfr = 1 THEN 1 ELSE 0 END) as ftfr_calls,
                SUM(attend_penalty) as total_attend_penalty,
                SUM(delay_penalty) as total_delay_penalty,
                SUM(total_penalty) as total_penalty,
                SUM(per_day_penalty) as total_per_day_penalty,
                AVG(CASE WHEN complaint_status = 'Final Closed' OR status = 'Closed' THEN total_downtime END) as avg_downtime_days,
                SUM(CASE WHEN attend_penalty > 0 THEN 1 ELSE 0 END) as attend_breach_count,
                SUM(CASE WHEN delay_penalty > 0 THEN 1 ELSE 0 END) as delay_breach_count,
                AVG(call_attend_hour_diff) / 24.0 as avg_attend_tat_days,
                AVG(total_downtime) / 24.0 as avg_close_tat_days
            FROM rj_penalties
            WHERE {where_str}
        """), params).fetchone()
        
        total_calls = totals[0] or 0
        closed_calls = totals[1] or 0
        ftfr_calls = totals[2] or 0
        total_attend = totals[3] or 0.0
        total_delay = totals[4] or 0.0
        total_net_penalty = totals[5] or 0.0
        total_per_day = totals[6] or 0.0
        avg_downtime = totals[7] or 0.0
        attend_breach_count = totals[8] or 0
        delay_breach_count = totals[9] or 0
        avg_attend_tat = totals[10] or 0.0
        avg_close_tat = totals[11] or 0.0
        
        ftfr_percentage = (ftfr_calls * 100.0 / closed_calls) if closed_calls > 0 else 0.0

        # 4. Daily complaint activity (Logged vs Closed)
        daily_logged = db.execute(text(f"""
            SELECT SUBSTR(complaint_raise_date, 1, 10) as day, COUNT(*) as count 
            FROM rj_penalties 
            WHERE {where_str} AND complaint_raise_date IS NOT NULL AND complaint_raise_date != ''
            GROUP BY day ORDER BY day DESC LIMIT 15
        """), params).fetchall()
        
        daily_closed = db.execute(text(f"""
            SELECT SUBSTR(complaint_close_date, 1, 10) as day, COUNT(*) as count 
            FROM rj_penalties 
            WHERE {where_str} AND complaint_close_date IS NOT NULL AND complaint_close_date != ''
            GROUP BY day ORDER BY day DESC LIMIT 15
        """), params).fetchall()

        # 5. Top Equipment Penalties
        equip_penalty = db.execute(text(f"""
            SELECT equipment_name, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND equipment_name IS NOT NULL AND equipment_name != ''
            GROUP BY equipment_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        # 6. Top District Penalties
        district_penalty = db.execute(text(f"""
            SELECT district_name, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND district_name IS NOT NULL AND district_name != ''
            GROUP BY district_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        # 7. Top Coordinator Penalties
        coord_penalty = db.execute(text(f"""
            SELECT coordinator_name, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND coordinator_name IS NOT NULL AND coordinator_name != ''
            GROUP BY coordinator_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        # 8. Zone breakdown
        zone_penalty = db.execute(text(f"""
            SELECT 
                CASE 
                    WHEN district_name IN ('Ajmer', 'Bhilwara', 'Nagaur', 'Tonk', 'Beawer', 'Kekri', 'Shahpura') THEN 'Ajmer'
                    WHEN district_name IN ('Jaipur', 'Alwar', 'Dausa', 'Jhunjhunu', 'Sikar', 'Dudu', 'Kotputli', 'Neem Ka Thana', 'Khairthal') THEN 'Jaipur'
                    WHEN district_name IN ('Jodhpur', 'Barmer', 'Jaisalmer', 'Jalore', 'Pali', 'Sirohi', 'Phalodi', 'Balotra', 'Sanchore') THEN 'Jodhpur'
                    WHEN district_name IN ('Bikaner', 'Churu', 'Hanumangarh', 'Sri Ganganagar', 'Ganganagar', 'Anupgarh') THEN 'Bikaner'
                    WHEN district_name IN ('Kota', 'Baran', 'Bundi', 'Jhalawar') THEN 'Kota'
                    WHEN district_name IN ('Udaipur', 'Banswara', 'Chittorgarh', 'Dungarpur', 'Rajsamand', 'Pratapgarh', 'Salumbar') THEN 'Udaipur'
                    ELSE 'Other'
                END as zone_name,
                SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND district_name IS NOT NULL AND district_name != ''
            GROUP BY zone_name
            ORDER BY total DESC
        """), params).fetchall()

        # 9. Hospital-wise Penalty Breakdown
        hospital_penalty = db.execute(text(f"""
            SELECT hospital_name, SUM(total_penalty) as total, COUNT(*) as count
            FROM rj_penalties
            WHERE {where_str} AND hospital_name IS NOT NULL AND hospital_name != ''
            GROUP BY hospital_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        # 10. Under Warranty vs Out of Warranty Penalty Share
        warranty_share = db.execute(text(f"""
            SELECT 
                CASE 
                    WHEN LOWER(is_under_warranty) LIKE '%yes%' OR is_under_warranty = '1' THEN 'Under Warranty'
                    ELSE 'Out of Warranty'
                END as warranty_status,
                SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str}
            GROUP BY warranty_status
        """), params).fetchall()

        # 11. Hospital Type Breakdown
        hosp_type_share = db.execute(text(f"""
            SELECT hospital_type, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND hospital_type IS NOT NULL AND hospital_type != ''
            GROUP BY hospital_type
            ORDER BY total DESC
        """), params).fetchall()

        # 12. Top Service Providers Penalties
        vendor_penalty = db.execute(text(f"""
            SELECT service_provider_name, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND service_provider_name IS NOT NULL AND service_provider_name != ''
            GROUP BY service_provider_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        # 13. Monthly Penalty Trend
        monthly_trend = db.execute(text(f"""
            SELECT month_text, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND month_text IS NOT NULL AND month_text != ''
            GROUP BY month_text
            ORDER BY MIN(id)
        """), params).fetchall()

        # 14. Monthly TAT Trend (New)
        monthly_tat = db.execute(text(f"""
            SELECT month_text, 
                   AVG(call_attend_hour_diff) / 24.0 as avg_attend_tat_days,
                   AVG(total_downtime) / 24.0 as avg_close_tat_days
            FROM rj_penalties
            WHERE {where_str} AND month_text IS NOT NULL AND month_text != ''
            GROUP BY month_text
            ORDER BY MIN(id)
        """), params).fetchall()

        # 15. Per Coordinator Workload Month-wise (New)
        coord_workload = db.execute(text(f"""
            SELECT coordinator_name, month_text, 
                   COUNT(*) as total_calls, 
                   SUM(CASE WHEN complaint_status = 'Final Closed' OR status = 'Closed' THEN 1 ELSE 0 END) as closed_calls
            FROM rj_penalties
            WHERE {where_str} AND coordinator_name IS NOT NULL AND coordinator_name != ''
            GROUP BY coordinator_name, month_text
            ORDER BY total_calls DESC
        """), params).fetchall()

        # 16. Daywise Penalty Breakdown (New)
        daywise_penalties = db.execute(text(f"""
            SELECT SUBSTR(complaint_raise_date, 1, 10) as day, 
                   SUM(attend_penalty) as attend_penalty, 
                   SUM(delay_penalty) as delay_penalty
            FROM rj_penalties
            WHERE {where_str} AND complaint_raise_date IS NOT NULL AND complaint_raise_date != ''
            GROUP BY day 
            ORDER BY day DESC 
            LIMIT 15
        """), params).fetchall()

        # 17. DI-wise Penalty Breakdown (New)
        di_penalty = db.execute(text(f"""
            SELECT di_name, SUM(total_penalty) as total
            FROM rj_penalties
            WHERE {where_str} AND di_name IS NOT NULL AND di_name != ''
            GROUP BY di_name
            ORDER BY total DESC LIMIT 8
        """), params).fetchall()

        res_data = {
            "success": True,
            "filter_options": {
                "districts": districts_list,
                "coordinators": coordinators_list,
                "zones": zones_list,
                "months": months_list,
                "equipments": equipments_list
            },
            "summary": {
                "total_calls": total_calls,
                "closed_calls": closed_calls,
                "ftfr_percentage": round(ftfr_percentage, 1),
                "total_attend_penalty": round(total_attend, 1),
                "total_delay_penalty": round(total_delay, 1),
                "total_penalty": round(total_net_penalty, 1),
                "total_per_day_penalty": round(total_per_day, 1),
                "avg_downtime_days": round(avg_downtime, 1),
                "attend_breach_count": attend_breach_count,
                "delay_breach_count": delay_breach_count,
                "avg_attend_tat_days": round(avg_attend_tat, 2),
                "avg_close_tat_days": round(avg_close_tat, 2)
            },
            "daily_activity": {
                "logged": [{"day": r[0], "count": r[1]} for r in daily_logged],
                "closed": [{"day": r[0], "count": r[1]} for r in daily_closed]
            },
            "breakdown": {
                "equipment": [{"name": r[0], "penalty": round(r[1], 1)} for r in equip_penalty],
                "district": [{"name": r[0], "penalty": round(r[1], 1)} for r in district_penalty],
                "coordinator": [{"name": r[0], "penalty": round(r[1], 1)} for r in coord_penalty],
                "zone": [{"name": r[0], "penalty": round(r[1], 1)} for r in zone_penalty],
                "hospital": [{"name": r[0], "penalty": round(r[1], 1), "count": r[2]} for r in hospital_penalty],
                "warranty": [{"status": r[0], "penalty": round(r[1], 1)} for r in warranty_share],
                "hospital_type": [{"type": r[0], "penalty": round(r[1], 1)} for r in hosp_type_share],
                "vendor": [{"name": r[0], "penalty": round(r[1], 1)} for r in vendor_penalty],
                "monthly_trend": [{"month": r[0], "penalty": round(r[1], 1)} for r in monthly_trend],
                "di": [{"name": r[0], "penalty": round(r[1], 1)} for r in di_penalty],
                "monthly_tat": [{"month": r[0], "avg_attend_tat_days": round(r[1] or 0.0, 2), "avg_close_tat_days": round(r[2] or 0.0, 2)} for r in monthly_tat],
                "coordinator_workload": [{"coordinator": r[0], "month": r[1], "total_calls": r[2], "closed_calls": r[3]} for r in coord_workload],
                "daywise_penalties": [{"day": r[0], "attend_penalty": round(r[1] or 0.0, 1), "delay_penalty": round(r[2] or 0.0, 1)} for r in daywise_penalties]
            }
        }
        cache.set(cache_key, res_data)
        return res_data
    except Exception as e:
        logger.error(f"Error fetching MIS dashboard metrics: {str(e)}")
        return {
            "success": False,
            "message": f"Server query error: {str(e)}"
        }

@router.post("/upload-penalties")
async def upload_excel_penalties(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Accepts the uploaded Rajasthan Penalty-Cyrix Excel sheet,
    parses it in memory, and bulk-inserts all rows into the rj_penalties table.
    """
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xlsm) are supported.")
        
    try:
        # Create table SQL
        create_table_sql = """
        CREATE TABLE IF NOT EXISTS rj_penalties (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sno TEXT,
            district_name TEXT,
            hospital_type TEXT,
            hospital_name TEXT,
            bar_code TEXT,
            equipment_name TEXT,
            equipment_model TEXT,
            complaint_id TEXT UNIQUE,
            complaint_raise_date TEXT,
            complaint_close_date TEXT,
            complaint_status TEXT,
            total_downtime REAL,
            estimated_cost REAL,
            penalty_days REAL,
            complaint_final_close TEXT,
            attend_date TEXT,
            attend_penalty REAL,
            delay_penalty REAL,
            total_penalty REAL,
            is_under_warranty TEXT,
            service_provider_name TEXT,
            status TEXT,
            equipment_type TEXT,
            asset_value REAL,
            complaint_logged_date TEXT,
            call_attend_hour_diff REAL,
            attented_per_day REAL,
            penalty_start_date TEXT,
            penalty_end_date TEXT,
            penalty_down_days REAL,
            penalty_slab REAL,
            penalty REAL,
            per_day_penalty REAL,
            total_penalty_calc REAL,
            total_per_day REAL,
            month_text TEXT,
            di_name TEXT,
            open_date TEXT,
            close_date TEXT,
            attend_delay_minutes REAL,
            same_day_close TEXT,
            standby TEXT,
            coordinator_name TEXT,
            final_close_month TEXT,
            close_month TEXT,
            eight_digit_code TEXT,
            open_days REAL,
            is_ftfr INTEGER DEFAULT 0
        );
        """
        # Ensure table and indexes exist
        db.execute(text(create_table_sql))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_district ON rj_penalties(district_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_coordinator ON rj_penalties(coordinator_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_equipment ON rj_penalties(equipment_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_month ON rj_penalties(month_text)"))
        db.commit()
        
        # Fetch existing complaint IDs to skip duplicate inserts
        existing_rows = db.execute(text("SELECT complaint_id FROM rj_penalties WHERE complaint_id IS NOT NULL AND complaint_id != ''")).fetchall()
        existing_ids = {r[0] for r in existing_rows}

        # Parse Excel from file stream
        contents = await file.read()
        import io
        from tempfile import NamedTemporaryFile
        
        # Write to temporary file so openpyxl read_only can open it
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name
            
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        if 'Penalty File' not in wb.sheetnames:
            os.unlink(tmp_path)
            raise HTTPException(status_code=400, detail="Sheet 'Penalty File' not found in workbook.")
            
        sheet = wb['Penalty File']
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter)

        def parse_date_to_iso(val):
            if not val:
                return None
            val = str(val).strip()
            formats = [
                "%d-%b-%Y %H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
                "%d-%b-%y %H:%M:%S",
                "%d-%b-%y",
                "%d-%b-%Y",
                "%Y-%m-%d"
            ]
            for fmt in formats:
                try:
                    return datetime.strptime(val, fmt)
                except Exception:
                    continue
            return None

        def safe_float(val):
            try:
                if val is None:
                    return 0.0
                return float(val)
            except Exception:
                return 0.0

        records = []
        row_count = 0
        
        for row in rows_iter:
            if row[0] is None:
                continue
                
            complaint_id = str(row[7]).strip() if row[7] is not None else ""
            if not complaint_id or complaint_id in existing_ids:
                continue
                
            raise_raw = row[8]
            close_raw = row[9]
            raise_dt = parse_date_to_iso(raise_raw)
            close_dt = parse_date_to_iso(close_raw)
            
            complaint_raise_date = raise_dt.strftime("%Y-%m-%d %H:%M:%S") if raise_dt else (str(raise_raw) if raise_raw else "")
            complaint_close_date = close_dt.strftime("%Y-%m-%d %H:%M:%S") if close_dt else (str(close_raw) if close_raw else "")
            
            is_ftfr = 0
            if raise_dt and close_dt:
                diff_hours = (close_dt - raise_dt).total_seconds() / 3600.0
                if diff_hours <= 24.0:
                    is_ftfr = 1

            rec = {
                "sno": str(row[0]),
                "district_name": str(row[1]) if row[1] is not None else "",
                "hospital_type": str(row[2]) if row[2] is not None else "",
                "hospital_name": str(row[3]) if row[3] is not None else "",
                "bar_code": str(row[4]) if row[4] is not None else "",
                "equipment_name": str(row[5]) if row[5] is not None else "",
                "equipment_model": str(row[6]) if row[6] is not None else "",
                "complaint_id": str(row[7]) if row[7] is not None else "",
                "complaint_raise_date": complaint_raise_date,
                "complaint_close_date": complaint_close_date,
                "complaint_status": str(row[10]) if row[10] is not None else "",
                "total_downtime": safe_float(row[11]),
                "estimated_cost": safe_float(row[12]),
                "penalty_days": safe_float(row[13]),
                "complaint_final_close": str(row[14]) if row[14] is not None else "",
                "attend_date": str(row[15]) if row[15] is not None else "",
                "attend_penalty": safe_float(row[16]),
                "delay_penalty": safe_float(row[17]),
                "total_penalty": safe_float(row[18]),
                "is_under_warranty": str(row[19]) if row[19] is not None else "",
                "service_provider_name": str(row[20]) if row[20] is not None else "",
                "status": str(row[21]) if row[21] is not None else "",
                "equipment_type": str(row[23]) if row[23] is not None else "",
                "asset_value": safe_float(row[24]),
                "complaint_logged_date": str(row[25]) if row[25] is not None else "",
                "call_attend_hour_diff": safe_float(row[26]),
                "attented_per_day": safe_float(row[28]) if len(row) > 28 else 0.0,
                "penalty_start_date": str(row[29]) if len(row) > 29 and row[29] is not None else "",
                "penalty_end_date": str(row[30]) if len(row) > 30 and row[30] is not None else "",
                "penalty_down_days": safe_float(row[31]) if len(row) > 31 else 0.0,
                "penalty_slab": safe_float(row[32]) if len(row) > 32 else 0.0,
                "penalty": safe_float(row[33]) if len(row) > 33 else 0.0,
                "per_day_penalty": safe_float(row[34]) if len(row) > 34 else 0.0,
                "total_penalty_calc": safe_float(row[35]) if len(row) > 35 else 0.0,
                "total_per_day": safe_float(row[36]) if len(row) > 36 else 0.0,
                "month_text": str(row[41]) if len(row) > 41 and row[41] is not None else "",
                "di_name": str(row[42]) if len(row) > 42 and row[42] is not None else "",
                "open_date": str(row[43]) if len(row) > 43 and row[43] is not None else "",
                "close_date": str(row[44]) if len(row) > 44 and row[44] is not None else "",
                "attend_delay_minutes": safe_float(row[45]) if len(row) > 45 else 0.0,
                "same_day_close": str(row[46]) if len(row) > 46 and row[46] is not None else "",
                "standby": str(row[48]) if len(row) > 48 and row[48] is not None else "",
                "coordinator_name": str(row[49]) if len(row) > 49 and row[49] is not None else "",
                "final_close_month": str(row[50]) if len(row) > 50 and row[50] is not None else "",
                "close_month": str(row[51]) if len(row) > 51 and row[51] is not None else "",
                "eight_digit_code": str(row[52]) if len(row) > 52 and row[52] is not None else "",
                "open_days": safe_float(row[53]) if len(row) > 53 else 0.0,
                "is_ftfr": is_ftfr
            }
            records.append(rec)
            row_count += 1
            
            # Flush every 500 records to keep memory low
            if len(records) >= 500:
                _bulk_insert(db, "rj_penalties", list(records[0].keys()), records)
                records = []
                
        if records:
            _bulk_insert(db, "rj_penalties", list(records[0].keys()), records)
            
        os.unlink(tmp_path)
        from app.utils import cache
        cache.clear_prefix("mis_dashboard:")
        return {
            "success": True,
            "message": f"Successfully parsed and synced {row_count} rows of Rajasthan Penalty Cyrix logs."
        }
    except Exception as e:
        logger.error(f"Error parsing uploaded Excel file: {str(e)}")
        return {
            "success": False,
            "message": f"Failed to upload: {str(e)}"
        }

@router.post("/upload-penalties-chunk")
async def upload_penalties_chunk(
    payload: dict,
    db: Session = Depends(get_db)
):
    """
    Accepts a JSON payload with pre-parsed penalty rows for chunked insertion.
    Payload: {"rows": [...list of row dicts...], "clear_first": true/false}
    This endpoint enables the client to split 46k rows into many small requests.
    """
    try:
        rows = payload.get("rows", [])
        clear_first = payload.get("clear_first", False)
        
        if clear_first:
            # Drop and create table with UNIQUE constraint
            db.execute(text("DROP TABLE IF EXISTS rj_penalties"))
            db.execute(text("""CREATE TABLE rj_penalties (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sno VARCHAR(20), district_name VARCHAR(200), hospital_type VARCHAR(100),
                hospital_name VARCHAR(255), bar_code VARCHAR(100), equipment_name VARCHAR(255),
                equipment_model VARCHAR(200), complaint_id VARCHAR(100) UNIQUE,
                complaint_raise_date VARCHAR(100), complaint_close_date VARCHAR(100),
                complaint_status VARCHAR(100), total_downtime FLOAT, estimated_cost FLOAT,
                penalty_days FLOAT, complaint_final_close VARCHAR(100), attend_date VARCHAR(100),
                attend_penalty FLOAT, delay_penalty FLOAT, total_penalty FLOAT,
                is_under_warranty VARCHAR(50), service_provider_name VARCHAR(200),
                status VARCHAR(100), equipment_type VARCHAR(100), asset_value FLOAT,
                complaint_logged_date VARCHAR(100), call_attend_hour_diff FLOAT,
                attented_per_day FLOAT, penalty_start_date VARCHAR(100),
                penalty_end_date VARCHAR(100), penalty_down_days FLOAT, penalty_slab FLOAT,
                penalty FLOAT, per_day_penalty FLOAT, total_penalty_calc FLOAT,
                total_per_day FLOAT, month_text VARCHAR(100), di_name VARCHAR(200),
                open_date VARCHAR(100), close_date VARCHAR(100), attend_delay_minutes FLOAT,
                same_day_close VARCHAR(50), standby VARCHAR(100),
                coordinator_name VARCHAR(200), final_close_month VARCHAR(100),
                close_month VARCHAR(100), eight_digit_code VARCHAR(100), open_days FLOAT,
                is_ftfr INTEGER DEFAULT 0
            )"""))
            db.commit()
        
        if rows:
            columns = list(rows[0].keys())
            _bulk_insert(db, "rj_penalties", columns, rows)
        
        from app.utils import cache
        cache.clear_prefix("mis_dashboard:")
        return {
            "success": True,
            "message": f"Inserted {len(rows)} rows. clear_first={clear_first}"
        }
    except Exception as e:
        logger.error(f"Error in penalties chunk upload: {str(e)}")
        return {
            "success": False,
            "message": f"Chunk upload failed: {str(e)}"
        }

@router.get("/existing-complaints")
async def get_existing_complaints(db: Session = Depends(get_db)):
    """
    Returns a list of all existing complaint_id values in the rj_penalties table
    to allow the client to filter duplicates before uploading.
    """
    try:
        # Check if table exists
        table_exists = db.execute(text(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='rj_penalties'"
        )).fetchone()
        
        if not table_exists:
            return {"success": True, "complaints": []}
            
        rows = db.execute(text("SELECT complaint_id FROM rj_penalties WHERE complaint_id IS NOT NULL AND complaint_id != ''")).fetchall()
        complaints = [r[0] for r in rows]
        return {"success": True, "complaints": complaints}
    except Exception as e:
        logger.error(f"Error fetching existing complaints: {str(e)}")
        return {"success": False, "message": str(e), "complaints": []}

@router.post("/upload-master-data")
async def upload_excel_master_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Accepts the uploaded Rajasthan Penalty-Cyrix Excel sheet,
    parses sheets ('DI Name List', 'Asset Value', 'Critical Equipment', 'Main Hospital'),
    and seeds respective tables (di_name_list, facility_details, asset_value_master, critical_equipment, main_hospitals).
    """
    if not file.filename.endswith(('.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx or .xlsm) are supported.")
        
    try:
        # Create tables SQL
        create_tables_sql = [
            """CREATE TABLE IF NOT EXISTS di_name_list (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                district_name VARCHAR(200),
                hospital_name VARCHAR(255),
                di_name VARCHAR(200),
                coordinator_name VARCHAR(200),
                zone_name VARCHAR(200)
            );""",
            """CREATE TABLE IF NOT EXISTS facility_details (
                facility_name VARCHAR(200) PRIMARY KEY,
                district_name VARCHAR(100),
                facility_incharge VARCHAR(100),
                dm_name VARCHAR(100),
                coordinator_name VARCHAR(100),
                facility_type VARCHAR(50),
                zone_name VARCHAR(50)
            );""",
            """CREATE TABLE IF NOT EXISTS asset_value_master (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_name VARCHAR(255) UNIQUE,
                rmsc_tender_cost FLOAT
            );""",
            """CREATE TABLE IF NOT EXISTS critical_equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_name VARCHAR(255) UNIQUE,
                classification VARCHAR(100)
            );""",
            """CREATE TABLE IF NOT EXISTS main_hospitals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                hospital_name VARCHAR(255) UNIQUE,
                hospital_type VARCHAR(200),
                sla_category VARCHAR(200)
            );"""
        ]
        
        for sql in create_tables_sql:
            db.execute(text(sql))
        db.commit()

        # Parse Excel from file stream
        contents = await file.read()
        from tempfile import NamedTemporaryFile
        import os
        
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(contents)
            tmp_path = tmp.name
            
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        
        # 1. Sync di_name_list and facility_details
        if 'DI Name List' in wb.sheetnames:
            sheet_di = wb['DI Name List']
            rows_di = sheet_di.iter_rows(values_only=True)
            next(rows_di) # skip header
            
            di_list_rows = []
            facility_details_map = {}
            for row in rows_di:
                if row[0] is None:
                    continue
                dist_name = str(row[0]).strip()
                hosp_name = str(row[1]).strip() if row[1] is not None else ""
                di_name = str(row[2]).strip() if row[2] is not None else ""
                coord_name = str(row[3]).strip() if row[3] is not None else ""
                zone_name = str(row[4]).strip() if row[4] is not None else ""
                
                di_list_rows.append({
                    "district_name": dist_name, "hospital_name": hosp_name,
                    "di_name": di_name, "coordinator_name": coord_name, "zone_name": zone_name
                })
                
                fac_type = "PHC"
                hosp_upper = hosp_name.upper()
                if "CHC" in hosp_upper: fac_type = "CHC"
                elif "DH" in hosp_upper or "DISTRICT HOSPITAL" in hosp_upper: fac_type = "DH"
                elif "SUB CENTER" in hosp_upper or "SUBCENTER" in hosp_upper or "SC" in hosp_upper: fac_type = "Sub-Center"
                elif "MEDICAL COLLEGE" in hosp_upper or "MEDICAL HOSPITAL" in hosp_upper: fac_type = "Medical College"
                
                if hosp_name and hosp_name not in facility_details_map:
                    facility_details_map[hosp_name] = {
                        "facility_name": hosp_name, "district_name": dist_name,
                        "facility_incharge": di_name, "dm_name": di_name, "coordinator_name": coord_name,
                        "facility_type": fac_type, "zone_name": zone_name
                    }
            
            # Insert DI Name List (batched multi-row inserts)
            db.execute(text("DELETE FROM di_name_list"))
            _bulk_insert(db, "di_name_list",
                         ["district_name", "hospital_name", "di_name", "coordinator_name", "zone_name"],
                         di_list_rows)
            
            # Insert Facility Details (batched multi-row inserts)
            db.execute(text("DELETE FROM facility_details"))
            _bulk_insert(db, "facility_details",
                         ["facility_name", "district_name", "facility_incharge", "dm_name", "coordinator_name", "facility_type", "zone_name"],
                         list(facility_details_map.values()))
            
        # 2. Sync asset_value_master
        if 'Asset Value' in wb.sheetnames:
            sheet_asset = wb['Asset Value']
            rows_asset = sheet_asset.iter_rows(values_only=True)
            next(rows_asset)
            
            seen_assets = set()
            asset_rows = []
            for row in rows_asset:
                if row[0] is None:
                    continue
                equip_name = str(row[0]).strip()
                try:
                    cost = float(row[1]) if row[1] is not None else 0.0
                except Exception:
                    cost = 0.0
                if equip_name not in seen_assets:
                    seen_assets.add(equip_name)
                    asset_rows.append({"equipment_name": equip_name, "rmsc_tender_cost": cost})
                    
            db.execute(text("DELETE FROM asset_value_master"))
            _bulk_insert(db, "asset_value_master", ["equipment_name", "rmsc_tender_cost"], asset_rows)
            
        # 3. Sync critical_equipment
        if 'Critical Equipment' in wb.sheetnames:
            sheet_crit = wb['Critical Equipment']
            rows_crit = sheet_crit.iter_rows(values_only=True)
            next(rows_crit)
            
            seen_crit = set()
            crit_rows = []
            for row in rows_crit:
                if row[0] is None:
                    continue
                equip_name = str(row[0]).strip()
                classification = str(row[1]).strip() if row[1] is not None else "Critical"
                if equip_name not in seen_crit:
                    seen_crit.add(equip_name)
                    crit_rows.append({"equipment_name": equip_name, "classification": classification})
                    
            db.execute(text("DELETE FROM critical_equipment"))
            _bulk_insert(db, "critical_equipment", ["equipment_name", "classification"], crit_rows)
            
        # 4. Sync main_hospitals
        if 'Main Hospital' in wb.sheetnames:
            sheet_main = wb['Main Hospital']
            rows_main = sheet_main.iter_rows(values_only=True)
            next(rows_main)
            
            seen_main = set()
            main_rows = []
            for row in rows_main:
                if row[0] is None:
                    continue
                hosp_name = str(row[0]).strip()
                hosp_type = str(row[1]).strip() if row[1] is not None else ""
                sla_category = str(row[4]).strip() if row[4] is not None else "MCH"
                if hosp_name not in seen_main:
                    seen_main.add(hosp_name)
                    main_rows.append({"hospital_name": hosp_name, "hospital_type": hosp_type, "sla_category": sla_category})
                    
            db.execute(text("DELETE FROM main_hospitals"))
            _bulk_insert(db, "main_hospitals", ["hospital_name", "hospital_type", "sla_category"], main_rows)
            
        os.unlink(tmp_path)
        return {
            "success": True,
            "message": "Successfully parsed and seeded all master data sheets (DI List, Assets, Main Hospitals)."
        }
    except Exception as e:
        logger.error(f"Error parsing master Excel sheets: {str(e)}")
        return {
            "success": False,
            "message": f"Failed to upload master data: {str(e)}"
        }

@router.post("/create-indexes")
async def create_indexes(db: Session = Depends(get_db)):
    """Creates indexes on rj_penalties table to optimize filtering queries."""
    try:
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_district ON rj_penalties(district_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_coordinator ON rj_penalties(coordinator_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_equipment ON rj_penalties(equipment_name)"))
        db.execute(text("CREATE INDEX IF NOT EXISTS idx_rj_penalties_month ON rj_penalties(month_text)"))
        db.commit()
        return {"success": True, "message": "Indexes created successfully on production D1 database."}
    except Exception as e:
        logger.error(f"Failed to create indexes: {str(e)}")
        return {"success": False, "message": str(e)}


# ======================== Asset Inventory Upload ========================

import csv as csv_module
import io as io_module
from datetime import datetime as dt_datetime

ASSETS_INVENTORY_TABLE = "assets_inventory"

# Table schema containing standard fields plus pre-parsed optimization columns
ASSETS_INVENTORY_CREATE_SQL = f"""CREATE TABLE IF NOT EXISTS {ASSETS_INVENTORY_TABLE} (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    district_name VARCHAR(200),
    hospital_name VARCHAR(300),
    department_name VARCHAR(300),
    group_name VARCHAR(300),
    equipment_name VARCHAR(300),
    model_name VARCHAR(300),
    serial_no VARCHAR(200),
    equipment_category VARCHAR(200),
    qr_code VARCHAR(200) UNIQUE,
    stock_register_page_no VARCHAR(100),
    received_date VARCHAR(100),
    installation_date VARCHAR(100),
    inventory_entry_date VARCHAR(100),
    moic_verified_date VARCHAR(100),
    po_date VARCHAR(100),
    po_cost VARCHAR(100),
    inventory_status VARCHAR(100),
    equipment_status VARCHAR(100),
    supplier VARCHAR(300),
    warranty_details VARCHAR(300),
    asset_value VARCHAR(100),
    di_name VARCHAR(200),
    dm_name VARCHAR(200),
    coordinator_name VARCHAR(200),
    zone_name VARCHAR(200),
    hospital_type VARCHAR(100),
    facility_type VARCHAR(100),
    equipment_type VARCHAR(200),
    
    -- Parsed columns for direct SQL query optimization
    is_verified INTEGER DEFAULT 0,
    warranty_expired INTEGER DEFAULT 1,
    parsed_asset_value REAL DEFAULT 0.0,
    moic_year INTEGER,
    moic_month INTEGER,
    install_year INTEGER,
    install_month INTEGER,
    
    uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
)"""

# 28 standard fields + 7 parsed helper fields
ASSETS_INVENTORY_COLUMNS = [
    "district_name", "hospital_name", "department_name", "group_name",
    "equipment_name", "model_name", "serial_no", "equipment_category",
    "qr_code", "stock_register_page_no", "received_date", "installation_date",
    "inventory_entry_date", "moic_verified_date", "po_date", "po_cost",
    "inventory_status", "equipment_status", "supplier", "warranty_details",
    "asset_value", "di_name", "dm_name", "coordinator_name", "zone_name",
    "hospital_type", "facility_type", "equipment_type",
    
    "is_verified", "warranty_expired", "parsed_asset_value",
    "moic_year", "moic_month", "install_year", "install_month"
]

CSV_HEADER_MAP = {
    "district name": "district_name",
    "hospital name": "hospital_name",
    "department name": "department_name",
    "group name": "group_name",
    "equipment name": "equipment_name",
    "model name": "model_name",
    "serial no": "serial_no",
    "serial no.": "serial_no",
    "equipment category": "equipment_category",
    "qr code": "qr_code",
    "stock register page no": "stock_register_page_no",
    "stock register page no.": "stock_register_page_no",
    "recieved date": "received_date",
    "received date": "received_date",
    "installation date": "installation_date",
    "inventory entry date": "inventory_entry_date",
    "moic verified date": "moic_verified_date",
    "po date": "po_date",
    "po cost": "po_cost",
    "inventory status": "inventory_status",
    "equipment status": "equipment_status",
    "supplier": "supplier",
    "warranty details": "warranty_details",
    "asset value": "asset_value",
    "di name": "di_name",
    "dm name": "dm_name",
    "coordinator name": "coordinator_name",
    "zone name": "zone_name",
    "hospital type": "hospital_type",
    "facility type": "facility_type",
    "equipment type": "equipment_type",
}



def _parse_date_flexible(date_str: str):

    """Try parsing various date formats and return a datetime object or None."""
    if not date_str or date_str.strip() in ("--", "", "NA", "N/A"):
        return None
    date_str = date_str.strip()
    formats = [
        "%d-%b-%Y",    # 17-May-2021
        "%d-%B-%Y",    # 17-May-2021
        "%Y-%m-%d",    # 2021-05-17
        "%d/%m/%Y",    # 17/05/2021
        "%m/%d/%Y",    # 05/17/2021
        "%d-%m-%Y",    # 17-05-2021
    ]
    for fmt in formats:
        try:
            return dt_datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def _is_warranty_expired(warranty_details: str) -> bool:
    """Check if warranty has expired based on 'start to end' format."""
    if not warranty_details or warranty_details.strip() in ("--", "", "NA", "N/A"):
        return True  # No warranty info = out of warranty
    parts = warranty_details.split(" to ")
    if len(parts) < 2:
        return True
    end_date = _parse_date_flexible(parts[-1].strip())
    if not end_date:
        return True
    return dt_datetime.now() > end_date


def _prepare_row(row: dict) -> dict:
    """Process a raw row dictionary, copies all 28 keys, and computes the 7 parsed optimization fields."""
    row_dict = {}
    
    # 1. Copy first 28 standard fields (or empty string if missing)
    for col in ASSETS_INVENTORY_COLUMNS[:28]:
        row_dict[col] = str(row.get(col, "")).strip()
        
    qr = str(row.get("qr_code", "")).strip()
    row_dict["qr_code"] = qr
    
    # 2. Parse asset value
    raw_val = row_dict["asset_value"]
    try:
        val = float(str(raw_val).replace(",", "").strip())
    except Exception:
        val = 0.0
    row_dict["parsed_asset_value"] = val
    
    # 3. Parse MOIC Verified Date & is_verified
    moic_date_str = row_dict["moic_verified_date"]
    moic_date = _parse_date_flexible(moic_date_str)
    if moic_date:
        row_dict["is_verified"] = 1
        row_dict["moic_year"] = moic_date.year
        row_dict["moic_month"] = moic_date.month
    else:
        row_dict["is_verified"] = 0
        row_dict["moic_year"] = None
        row_dict["moic_month"] = None
        
    # 4. Parse Installation Date
    install_date_str = row_dict["installation_date"]
    install_date = _parse_date_flexible(install_date_str)
    if install_date:
        row_dict["install_year"] = install_date.year
        row_dict["install_month"] = install_date.month
    else:
        row_dict["install_year"] = None
        row_dict["install_month"] = None
        
    # 5. Parse Warranty Details
    warranty_str = row_dict["warranty_details"]
    warranty_expired = _is_warranty_expired(warranty_str)
    row_dict["warranty_expired"] = 1 if warranty_expired else 0
    
    return row_dict


@router.post("/upload-assets-chunk")
async def upload_assets_chunk(
    payload: dict,
    db: Session = Depends(get_db)
):
    """Accepts a JSON payload with pre-parsed asset inventory rows for chunked insertion."""
    try:
        rows = payload.get("rows", [])
        clear_first = payload.get("clear_first", False)

        if clear_first:
            db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
            db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
            db.commit()

        valid_rows = []
        skipped = 0
        for row in rows:
            qr = str(row.get("qr_code", "")).strip()
            if not qr or qr == "--":
                skipped += 1
                continue

            prepared = _prepare_row(row)
            valid_rows.append(prepared)

        if valid_rows:
            _bulk_insert(db, ASSETS_INVENTORY_TABLE, ASSETS_INVENTORY_COLUMNS, valid_rows)

        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_district ON {ASSETS_INVENTORY_TABLE}(district_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_qr ON {ASSETS_INVENTORY_TABLE}(qr_code)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_hospital ON {ASSETS_INVENTORY_TABLE}(hospital_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_zone ON {ASSETS_INVENTORY_TABLE}(zone_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_di ON {ASSETS_INVENTORY_TABLE}(di_name)"))
        db.commit()

        return {
            "success": True,
            "inserted": len(valid_rows),
            "skipped": skipped,
            "message": f"Successfully inserted {len(valid_rows)} rows"
        }
    except Exception as e:
        logger.error(f"Error in upload_assets_chunk: {str(e)}")
        return {
            "success": False,
            "inserted": 0,
            "skipped": 0,
            "message": f"Asset upload failed: {str(e)}"
        }


@router.post("/upload-assets-bulk")
async def upload_assets_bulk(
    payload: dict,
    db: Session = Depends(get_db)
):
    """Accepts a JSON payload with pre-parsed asset inventory rows from the frontend and replaces all data."""
    import time
    start_time = time.perf_counter()
    try:
        rows = payload.get("rows", [])
        skipped_on_client = payload.get("skipped_on_client", 0)

        valid_rows = []
        skipped = 0
        for row in rows:
            qr = str(row.get("qr_code", "")).strip()
            if not qr or qr == "--":
                skipped += 1
                continue

            prepared = _prepare_row(row)
            valid_rows.append(prepared)

        # Drop and recreate table
        db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
        db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
        db.commit()

        if valid_rows:
            _bulk_insert(db, ASSETS_INVENTORY_TABLE, ASSETS_INVENTORY_COLUMNS, valid_rows)

        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_district ON {ASSETS_INVENTORY_TABLE}(district_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_qr ON {ASSETS_INVENTORY_TABLE}(qr_code)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_hospital ON {ASSETS_INVENTORY_TABLE}(hospital_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_zone ON {ASSETS_INVENTORY_TABLE}(zone_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_di ON {ASSETS_INVENTORY_TABLE}(di_name)"))
        db.commit()

        elapsed_ms = (time.perf_counter() - start_time) * 1000

        return {
            "success": True,
            "inserted": len(valid_rows),
            "skipped": skipped,
            "elapsed_ms": round(elapsed_ms, 2),
            "message": f"Successfully inserted {len(valid_rows)} rows in {elapsed_ms:.1f}ms"
        }
    except Exception as e:
        logger.error(f"Error in JSON asset upload: {str(e)}")
        return {
            "success": False,
            "inserted": 0,
            "skipped": 0,
            "message": f"Asset upload failed: {str(e)}"
        }


@router.post("/upload-assets-csv")
async def upload_assets_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Processes asset inventory import directly from a file upload."""
    import time
    start_time = time.perf_counter()
    try:
        contents = await file.read()
        try:
            text_content = contents.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text_content = contents.decode("utf-8", errors="replace")
            except UnicodeDecodeError:
                text_content = contents.decode("latin-1")

        first_line = text_content.split("\n", 1)[0]
        delimiter = "\t" if "\t" in first_line else ","

        reader = csv_module.reader(io_module.StringIO(text_content), delimiter=delimiter)
        rows_iter = iter(reader)

        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {"success": False, "inserted": 0, "skipped": 0, "message": "CSV file is empty."}

        col_map = []
        for idx, h in enumerate(header_row):
            normalized = h.strip().strip('"').lower().strip()
            api_key = CSV_HEADER_MAP.get(normalized)
            if api_key:
                col_map.append((idx, api_key))

        if not col_map:
            return {"success": False, "inserted": 0, "skipped": 0,
                    "message": "Could not match any CSV headers."}

        valid_rows = []
        skipped = 0
        for parts in rows_iter:
            if not parts or not any(p.strip() for p in parts):
                continue
            row = {}
            for csv_idx, api_key in col_map:
                row[api_key] = parts[csv_idx].strip().strip('"') if csv_idx < len(parts) else ""

            qr = row.get("qr_code", "").strip()
            if not qr or qr == "--":
                skipped += 1
                continue
            row["qr_code"] = qr

            # Map the parsed helper attributes
            prepared = _prepare_row(row)
            valid_rows.append(prepared)

        if not valid_rows:
            return {"success": False, "inserted": 0, "skipped": skipped,
                    "message": "No valid rows found."}

        db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
        db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
        db.commit()

        _bulk_insert(db, ASSETS_INVENTORY_TABLE, ASSETS_INVENTORY_COLUMNS, valid_rows)

        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_district ON {ASSETS_INVENTORY_TABLE}(district_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_qr ON {ASSETS_INVENTORY_TABLE}(qr_code)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_hospital ON {ASSETS_INVENTORY_TABLE}(hospital_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_zone ON {ASSETS_INVENTORY_TABLE}(zone_name)"))
        db.execute(text(f"CREATE INDEX IF NOT EXISTS idx_assets_inv_di ON {ASSETS_INVENTORY_TABLE}(di_name)"))
        db.commit()

        elapsed_ms = (time.perf_counter() - start_time) * 1000

        return {
            "success": True,
            "inserted": len(valid_rows),
            "skipped": skipped,
            "elapsed_ms": round(elapsed_ms, 2),
            "message": f"Inserted {len(valid_rows)} rows in {elapsed_ms:.1f}ms"
        }
    except Exception as e:
        logger.error(f"Error in CSV asset upload: {str(e)}")
        return {
            "success": False,
            "inserted": 0,
            "skipped": 0,
            "message": f"Asset upload failed: {str(e)}"
        }


@router.get("/assets-inventory")
async def get_assets_inventory(
    district: str = None,
    hospital: str = None,
    zone: str = None,
    di: str = None,
    month: str = None, # format: "YYYY-MM"
    equipment_status: str = None,
    search: str = None,
    page: int = 1,
    page_size: int = 100,
    db: Session = Depends(get_db)
):
    """Get paginated asset inventory with optional filters served directly using optimized SQLite queries."""
    try:
        # Check if table schema has is_verified columns
        try:
            db.execute(text(f"SELECT is_verified FROM {ASSETS_INVENTORY_TABLE} LIMIT 1")).fetchone()
        except Exception:
            db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
            db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
            db.commit()

        where_clauses = []
        params = {}

        if district:
            where_clauses.append("district_name = :district")
            params["district"] = district
        if hospital:
            where_clauses.append("hospital_name = :hospital")
            params["hospital"] = hospital
        if zone:
            where_clauses.append("zone_name = :zone")
            params["zone"] = zone
        if di:
            where_clauses.append("di_name = :di")
            params["di"] = di
        if equipment_status:
            where_clauses.append("equipment_status = :equipment_status")
            params["equipment_status"] = equipment_status

        if month:
            try:
                parts = month.split("-")
                target_year = int(parts[0])
                target_month = int(parts[1])
                where_clauses.append("is_verified = 1 AND moic_year = :target_year AND moic_month = :target_month")
                params["target_year"] = target_year
                params["target_month"] = target_month
            except Exception:
                pass

        if search:
            search_pattern = f"%{search.strip()}%"
            where_clauses.append(
                "(equipment_name LIKE :search OR qr_code LIKE :search OR serial_no LIKE :search OR hospital_name LIKE :search)"
            )
            params["search"] = search_pattern

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # 1. Total row count query
        count_sql = f"SELECT COUNT(*) FROM {ASSETS_INVENTORY_TABLE} WHERE {where_sql}"
        total = db.execute(text(count_sql), params).scalar() or 0

        # 2. Paginated data query
        offset = (page - 1) * page_size
        data_sql = f"""
        SELECT * FROM {ASSETS_INVENTORY_TABLE} 
        WHERE {where_sql} 
        ORDER BY id DESC 
        LIMIT :limit OFFSET :offset
        """
        rows = db.execute(text(data_sql), {**params, "limit": page_size, "offset": offset}).fetchall()

        columns = ["id"] + ASSETS_INVENTORY_COLUMNS + ["uploaded_at"]
        assets = []
        for r in rows:
            row_dict = {}
            for i, col in enumerate(columns):
                row_dict[col] = r[i] if i < len(r) else None
            assets.append(row_dict)

        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "assets": assets
        }
    except Exception as e:
        logger.error(f"Error fetching assets inventory: {str(e)}")
        return {"success": False, "total": 0, "assets": [], "message": str(e)}


@router.get("/assets-filters")
async def get_assets_filters(db: Session = Depends(get_db)):
    """Get distinct filter lists and combinations for dependent dropdowns from D1."""
    try:
        try:
            db.execute(text(f"SELECT is_verified FROM {ASSETS_INVENTORY_TABLE} LIMIT 1")).fetchone()
        except Exception:
            db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
            db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
            db.commit()

        comb_sql = f"""
        SELECT DISTINCT zone_name, district_name, di_name 
        FROM {ASSETS_INVENTORY_TABLE} 
        WHERE zone_name IS NOT NULL AND zone_name != ''
        """
        comb_rows = db.execute(text(comb_sql)).fetchall()

        valid_rajasthan_zones = {"Ajmer", "Bikaner", "Jaipur", "Jodhpur", "Kota", "Udaipur", "Bharatpur"}
        combinations = []
        zones_set = set()
        districts_set = set()
        di_names_set = set()

        for z, d, di in comb_rows:
            z_clean = str(z or "").strip()
            matched_zone = None
            for rz in valid_rajasthan_zones:
                if rz.lower() in z_clean.lower():
                    matched_zone = rz
                    break

            if matched_zone:
                zones_set.add(matched_zone)
                districts_set.add(str(d or "").strip())
                di_names_set.add(str(di or "").strip())
                combinations.append({
                    "zone": matched_zone,
                    "district": str(d or "").strip(),
                    "di": str(di or "").strip()
                })

        month_sql = f"""
        SELECT DISTINCT moic_year, moic_month 
        FROM {ASSETS_INVENTORY_TABLE} 
        WHERE is_verified = 1 AND moic_year IS NOT NULL AND moic_month IS NOT NULL
        ORDER BY moic_year DESC, moic_month DESC
        """
        month_rows = db.execute(text(month_sql)).fetchall()
        months = [f"{r[0]}-{str(r[1]).zfill(2)}" for r in month_rows]

        return {
            "success": True,
            "zones": sorted(list(zones_set)),
            "districts": sorted(list(districts_set)),
            "di_names": sorted(list(di_names_set)),
            "months": months,
            "combinations": combinations
        }
    except Exception as e:
        logger.error(f"Error fetching assets filters: {str(e)}")
        return {"success": False, "zones": [], "districts": [], "di_names": [], "months": [], "combinations": []}


@router.get("/assets-stats")
async def get_assets_stats(
    zone: str = None,
    district: str = None,
    di: str = None,
    month: str = None, # format: "YYYY-MM"
    db: Session = Depends(get_db)
):
    """Get MIS dashboard summary metrics and chart items computed directly via SQL aggregation (0 RAM / Timeout risk)."""
    try:
        try:
            db.execute(text(f"SELECT is_verified FROM {ASSETS_INVENTORY_TABLE} LIMIT 1")).fetchone()
        except Exception:
            db.execute(text(f"DROP TABLE IF EXISTS {ASSETS_INVENTORY_TABLE}"))
            db.execute(text(ASSETS_INVENTORY_CREATE_SQL))
            db.commit()

        where_clauses = []
        params = {}

        if zone:
            where_clauses.append("zone_name = :zone")
            params["zone"] = zone
        if district:
            where_clauses.append("district_name = :district")
            params["district"] = district
        if di:
            where_clauses.append("di_name = :di")
            params["di"] = di

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # 1. SQL Aggregations
        agg_sql = f"""
        SELECT 
            COUNT(*) as total_equipment,
            SUM(is_verified) as verified_equipment,
            SUM(CASE WHEN warranty_expired = 0 THEN 1 ELSE 0 END) as under_warranty,
            SUM(warranty_expired) as out_of_warranty,
            SUM(parsed_asset_value) as total_value,
            SUM(CASE WHEN is_verified = 1 THEN parsed_asset_value ELSE 0 END) as verified_value,
            SUM(CASE WHEN is_verified = 1 AND warranty_expired = 1 THEN parsed_asset_value ELSE 0 END) as verified_out_of_warranty_value
        FROM {ASSETS_INVENTORY_TABLE}
        WHERE {where_sql}
        """
        agg_res = db.execute(text(agg_sql), params).fetchone()

        total_equipment = agg_res[0] or 0
        verified_count = agg_res[1] or 0
        under_warranty_count = agg_res[2] or 0
        out_of_warranty_count = agg_res[3] or 0
        total_value = agg_res[4] or 0.0
        verified_value = agg_res[5] or 0.0
        verified_out_of_warranty_value = agg_res[6] or 0.0

        # 2. Arrear billing calculation
        now = dt_datetime.now()
        if month:
            try:
                parts = month.split("-")
                target_year = int(parts[0])
                target_month = int(parts[1])
            except Exception:
                target_year = now.year
                target_month = now.month
        else:
            target_year = now.year
            target_month = now.month

        # Calculate elapsed months for verified devices in the target month
        arrear_sql = f"""
        SELECT 
            parsed_asset_value, install_year, install_month
        FROM {ASSETS_INVENTORY_TABLE}
        WHERE is_verified = 1 
          AND moic_year = :target_year 
          AND moic_month = :target_month
          AND {where_sql}
        """
        arrear_rows = db.execute(text(arrear_sql), {**params, "target_year": target_year, "target_month": target_month}).fetchall()
        arrear_billing = 0.0
        for r_val, i_yr, i_mo in arrear_rows:
            if r_val and i_yr and i_mo:
                monthly_rate = (r_val * 6.08 / 100) / 12
                months_diff = (target_year - i_yr) * 12 + (target_month - i_mo)
                if months_diff > 0:
                    arrear_billing += monthly_rate * months_diff

        monthly_value = (verified_out_of_warranty_value * 6.08 / 100) / 12
        total_billing = monthly_value + arrear_billing

        # 3. Chart 1: Status Distribution
        status_sql = f"""
        SELECT equipment_status, COUNT(*) 
        FROM {ASSETS_INVENTORY_TABLE} 
        WHERE {where_sql} 
        GROUP BY equipment_status
        """
        status_rows = db.execute(text(status_sql), params).fetchall()
        status_list = [{"name": r[0] or "Unknown", "value": r[1]} for r in status_rows]

        # 4. Chart 2: Top 5 Types
        type_sql = f"""
        SELECT equipment_type, COUNT(*) as cnt 
        FROM {ASSETS_INVENTORY_TABLE} 
        WHERE {where_sql} 
        GROUP BY equipment_type 
        ORDER BY cnt DESC 
        LIMIT 5
        """
        type_rows = db.execute(text(type_sql), params).fetchall()
        top_types = [{"name": r[0] or "Other", "value": r[1]} for r in type_rows]

        # 5. Chart 3: Warranty Breakdown
        warranty_sql = f"""
        SELECT warranty_expired, COUNT(*) 
        FROM {ASSETS_INVENTORY_TABLE} 
        WHERE {where_sql} 
        GROUP BY warranty_expired
        """
        warranty_rows = db.execute(text(warranty_sql), params).fetchall()
        warranty_dict = {0: 0, 1: 0}
        for r in warranty_rows:
            warranty_dict[int(r[0] or 0)] = r[1]
        warranty_list = [
            {"name": "Under Warranty", "value": warranty_dict[0]},
            {"name": "Out of Warranty", "value": warranty_dict[1]}
        ]

        return {
            "success": True,
            "total_equipment": total_equipment,
            "verified_equipment": verified_count,
            "under_warranty": under_warranty_count,
            "out_of_warranty": out_of_warranty_count,
            "total_value": round(total_value, 2),
            "verified_value": round(verified_value, 2),
            "verified_out_of_warranty_value": round(verified_out_of_warranty_value, 2),
            "monthly_value": round(monthly_value, 2),
            "arrear_billing": round(arrear_billing, 2),
            "total_billing": round(total_billing, 2),
            "charts": {
                "top_types": top_types,
                "status_list": status_list,
                "warranty_list": warranty_list
            }
        }
    except Exception as e:
        logger.error(f"Error fetching assets stats: {str(e)}")
        return {
            "success": False, "total_equipment": 0, "verified_equipment": 0,
            "under_warranty": 0, "out_of_warranty": 0, "total_value": 0,
            "verified_value": 0, "verified_out_of_warranty_value": 0,
            "monthly_value": 0, "arrear_billing": 0, "total_billing": 0,
            "charts": {"top_types": [], "status_list": [], "warranty_list": []}
        }



